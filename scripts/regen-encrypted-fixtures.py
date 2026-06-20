"""
Regenerate tests/fixtures/encrypted_sample.{pdf,xlsx} from a clean PalmPay
PDF + a synthetic XLSX. Password: "test123".

Owner-only tool. Not invoked in CI, not a dev-dep. Run when:
  - pikepdf / msoffcrypto-tool ship a behavior change we want to track
  - a new boundary test needs a different password

Requires:
  uv add --dev pikepdf msoffcrypto-tool   (already in dev deps for tests)

Usage:
  uv run python scripts/regen-encrypted-fixtures.py
"""

from __future__ import annotations

import subprocess
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
FIXTURES = ROOT / "tests" / "fixtures"
ENC_PDF = FIXTURES / "encrypted_sample.pdf"
ENC_XLSX = FIXTURES / "encrypted_sample.xlsx"
PASSWORD = "test123"


def regen_pdf() -> None:
    """Build a minimal one-page PDF inline (no real content, ~2KB), then
    encrypt it. Source is synthetic to keep the fixture tiny + obviously
    fake — no risk of stray PII surviving the round-trip from a real bank
    statement."""
    import pikepdf

    pdf = pikepdf.new()
    pdf.add_blank_page(page_size=(595.0, 842.0))  # A4
    pdf.save(
        ENC_PDF,
        encryption=pikepdf.Encryption(user=PASSWORD, owner=PASSWORD),
    )
    print(f"wrote {ENC_PDF} ({ENC_PDF.stat().st_size} bytes)")


def regen_xlsx() -> None:
    from openpyxl import Workbook

    plain = FIXTURES / "_plain_for_encrypt.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Test User"
    ws["A2"] = "Test row"
    wb.save(plain)
    try:
        subprocess.run(
            [
                "msoffcrypto-tool",
                "-e",
                "-p",
                PASSWORD,
                str(plain),
                str(ENC_XLSX),
            ],
            check=True,
            capture_output=True,
        )
    finally:
        plain.unlink(missing_ok=True)
    print(f"wrote {ENC_XLSX} ({ENC_XLSX.stat().st_size} bytes)")


if __name__ == "__main__":
    FIXTURES.mkdir(parents=True, exist_ok=True)
    regen_pdf()
    regen_xlsx()
