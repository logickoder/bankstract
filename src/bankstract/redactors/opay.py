"""
OPay statement redactor.

OPay narration is the largest leak surface: every Transfer line carries the
counterparty's full name + bank + partial account number. A vocabulary
keep-list would leak somewhere, so — like FBN and Zenith — the safe move is
to blank the entire description column wholesale.

Strategy:
1. Header: the row directly beneath "Account Name Account Number Address"
   carries the holder name, account number, and street address as a single
   row. Blank it, restore obviously-fake placeholders.
2. Body: every word inside the description (175..300) or reference
   (475..560) x0 range is blanked. The channel column ("Mobile") is
   structural and stays. Dates, amounts, and the balance column survive so
   parser tests can still assert totals + counts.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

from .._layout import classify
from .._pymupdf import rect as _rect
from .._source import Source, rewind
from .._xlsx import sniff_format
from ..schema import RedactReport, RedactResult
from . import register
from ._shared import (
    DEFAULT_SWEEPS,
    apply_regex_sweeps,
    page_rows,
    redact_rect,
    redact_word,
    shape_preserve,
)
from .base import Redactor

# Column ranges are deliberately WIDER than the parser's COL_* — the parser
# wants precise column attribution, the redactor wants to catch every word
# that could be narration regardless of layout drift between statements.
# Section 1 (Wallet Account) and section 2 (Savings Account) use different
# left-shifts; the unions cover both. The channel token ("Mobile") at
# x0~395-475 has classify()=="text" so it never trips the alnum check.
DESC_ZONE: tuple[float, float] = (170.0, 305.0)
REF_ZONE: tuple[float, float] = (450.0, 565.0)

ROW_TOL = 4.0

FORMAT_VERSION_PDF = "opay-pdf-2026-01"
FORMAT_VERSION_XLSX = "opay-xlsx-2026-01"


_XLSX_HEADER_REPLACEMENTS: dict[int, object] = {
    # Row-1 columns (0-indexed): holder name (1), account number (3), address (5).
    1: "TEST USER",
    3: "0000000000",
    5: "Test Address",
}


def _xlsx_redact_sheet(ws: Any, audit: list[str]) -> int:
    """Blank PII cells in one OPay-shape worksheet. Returns redaction count."""
    n = 0
    # Header row (row 2 in 1-indexed openpyxl; rows[1] in 0-indexed iter).
    for col_idx, replacement in _XLSX_HEADER_REPLACEMENTS.items():
        cell = ws.cell(row=2, column=col_idx + 1)
        if cell.value is not None:
            audit.append(f"header[r2c{col_idx + 1}]: {cell.value!r} -> {replacement!r}")
            cell.value = replacement
            n += 1
    # Tx rows start at row 8 (after rows 1-5 metadata, 6 blank, 7 column header).
    row = 8
    while True:
        date_cell = ws.cell(row=row, column=1)
        if date_cell.value is None or str(date_cell.value).strip() == "":
            break
        # Column 3 = Description, column 8 = Transaction Reference.
        desc = ws.cell(row=row, column=3)
        if desc.value is not None:
            audit.append(f"desc[r{row}]: {str(desc.value)[:30]!r} -> ''")
            desc.value = ""
            n += 1
        ref = ws.cell(row=row, column=8)
        if ref.value is not None:
            original = str(ref.value)
            shape = shape_preserve(original)
            audit.append(f"ref[r{row}]: {original!r} -> {shape!r}")
            ref.value = shape
            n += 1
        row += 1
    return n


def _redact_xlsx(source: Source) -> RedactResult:
    rewind(source)
    handle: Any = source if hasattr(source, "read") else str(source)
    wb = load_workbook(handle)
    report = RedactReport(bank="opay")
    try:
        for sheet_name in wb.sheetnames:
            audit: list[str] = []
            n = _xlsx_redact_sheet(wb[sheet_name], audit)
            report.pages += 1
            report.redactions += n
            report.audit.append((report.pages, audit))
        buf = BytesIO()
        wb.save(buf)
        data = buf.getvalue()
    finally:
        wb.close()
    return RedactResult(
        data=data,
        bank="opay",
        format="xlsx",
        format_version=FORMAT_VERSION_XLSX,
        report=report,
    )


class OPayRedactor(Redactor):
    bank = "opay"
    supported_formats = ("pdf", "xlsx")
    format_version = FORMAT_VERSION_PDF

    def redact(self, source: Source) -> RedactResult:
        # Format dispatch by extension. XLSX path uses openpyxl cell-level
        # rewrite (no font / layout drift to worry about); PDF path falls
        # through to the inherited template-method pipeline.
        try:
            fmt = sniff_format(source)
        except ValueError:
            return super().redact(source)
        if fmt == "xlsx":
            return _redact_xlsx(source)
        return super().redact(source)

    def redact_header(
        self,
        page: Any,
        pending_text: list[tuple[Any, str]],
        audit: list[str],
    ) -> None:
        # Account holder + account number + address sit on the single line
        # directly below the "Account Name Account Number Address" label
        # row. Blank that whole strip and restore fixed placeholders for
        # each column.
        for hit in page.search_for("Account Name"):
            row_top = hit.y1 + 2
            row_bottom = hit.y1 + 14
            holder_rect = _rect(60.0, row_top, 220.0, row_bottom)
            acct_rect = _rect(220.0, row_top, 380.0, row_bottom)
            addr_rect = _rect(380.0, row_top, page.rect.x1 - 20, row_bottom)
            for r, text, label in (
                (holder_rect, "TEST USER", "Account Name"),
                (acct_rect, "0000000000", "Account Number"),
                (addr_rect, "Test Address", "Address"),
            ):
                redact_rect(page, r, text, pending_text)
                audit.append(f"header[{label}] -> {text!r}")

        # "Generated on DD Mon YYYY HH:MM:SS" reveals print time; replace
        # with a fixed placeholder so two runs produce identical fixtures.
        for hit in page.search_for("Generated on"):
            line = _rect(hit.x1 + 2, hit.y0, page.rect.x1 - 20, hit.y1 + 2)
            redact_rect(page, line, "01 Jan 2026 00:00:00", pending_text)
            audit.append("header[Generated on] -> placeholder")

    def redact_body(
        self,
        page: Any,
        pending_text: list[tuple[Any, str]],
        audit: list[str],
    ) -> None:
        rows = page_rows(page, ROW_TOL)

        # Page 1 alone has chrome above the body (Wallet Account header,
        # holder block, period line, balance summary). Anchor on the first
        # "Trans." column header on page 1 to skip past it. On later pages
        # we default body_y_min=0 so the entire page is body; the section-2
        # break on page 35 repeats the column header MID-PAGE, but its
        # surrounding chrome words (Savings/Account/Period:/₦-summary)
        # mostly fall outside the desc + ref x-zones — and anchoring on
        # that mid-page header would skip every section-1 tx above it,
        # leaving their refs un-redacted.
        body_y_min: float = 0.0
        if page.number == 0:
            for row in rows:
                if any(w.text == "Trans." for w in row):
                    body_y_min = max(w.bottom for w in row)
                    break

        for row in rows:
            covered: set[int] = set()
            apply_regex_sweeps(page, row, DEFAULT_SWEEPS, pending_text, audit, covered)

            row_top = row[0].top
            if row_top < body_y_min:
                continue
            for idx, w in enumerate(row):
                if idx in covered:
                    continue
                kind = classify(w.text)
                in_desc = DESC_ZONE[0] <= w.x0 < DESC_ZONE[1]
                in_ref = REF_ZONE[0] <= w.x0 < REF_ZONE[1]
                # Preserve "--" debit/credit placeholders even though they
                # sit in the desc x-range on section 2 (which uses a
                # left-shifted layout). Their classify() result is "text"
                # but the literal token has structural meaning the parser
                # needs to round-trip.
                if in_desc and kind in ("text", "alnum") and w.text != "--":
                    redact_word(page, w, "", pending_text)
                    covered.add(idx)
                    audit.append(f"desc: {w.text!r} -> (blank)")
                elif in_ref and kind == "alnum":
                    # Shape-preserve so the parser's alnum classifier still
                    # recognises it after redaction (FBN pattern).
                    shape = shape_preserve(w.text)
                    redact_word(page, w, shape, pending_text)
                    covered.add(idx)
                    audit.append(f"ref: {w.text!r} -> {shape!r}")


register(OPayRedactor())
